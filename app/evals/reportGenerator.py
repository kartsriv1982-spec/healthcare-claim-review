
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment


class ReportGenerator:

    def __init__(self,
                 output_folder="evals/reports"):

        self.output_folder = Path(output_folder)

        self.output_folder.mkdir(
            parents=True,
            exist_ok=True
        )

    # ------------------------------------------------------

    def generate(
        self,
        evaluations
    ):

        wb = Workbook()

        summary_sheet = wb.active
        summary_sheet.title = "Summary"

        self.build_summary(
            summary_sheet,
            evaluations
        )

        overall_sheet = wb.create_sheet(
            "Overall Scores"
        )

        self.build_overall_scores(
            overall_sheet,
            evaluations
        )

        metric_sheet = wb.create_sheet(
            "Metric Details"
        )

        self.build_metric_details(
            metric_sheet,
            evaluations
        )

        claim_sheet = wb.create_sheet(
            "Claim Details"
        )

        self.build_claim_details(
            claim_sheet,
            evaluations
        )

        filename = self.output_folder / \
            f"evaluation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        wb.save(filename)

        print(f"Report saved to {filename}")

        return filename

    # ------------------------------------------------------

    def style_header(
        self,
        ws,
        row=1
    ):

        fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78"
        )

        for cell in ws[row]:

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = fill

            cell.alignment = Alignment(
                horizontal="center"
            )

    # ------------------------------------------------------

    def build_summary(
        self,
        ws,
        evaluations
    ):

        ws.append(
            ["Metric", "Value"]
        )

        self.style_header(ws)

        total = len(evaluations)

        passed = sum(
            1
            for e in evaluations
            if e["overall"].passed
        )

        failed = total - passed

        average = round(

            sum(

                e["overall"].score

                for e in evaluations

            ) / total,

            3

        )

        ws.append(["Run Date", str(datetime.now())])
        ws.append(["Claims Evaluated", total])
        ws.append(["Passed", passed])
        ws.append(["Failed", failed])
        ws.append(["Average Score", average])

    # ------------------------------------------------------

    def build_overall_scores(
        self,
        ws,
        evaluations
    ):

        ws.append([
            "Claim ID",
            "Overall Score",
            "Passed"
        ])

        self.style_header(ws)

        for e in evaluations:

            ws.append([

                e["claim_id"],

                e["overall"].score,

                e["overall"].passed

            ])

    # ------------------------------------------------------

    def build_metric_details(
        self,
        ws,
        evaluations
    ):

        ws.append([

            "Claim ID",

            "Metric",

            "Score",

            "Threshold",

            "Passed",

            "Remarks"

        ])

        self.style_header(ws)

        for evaluation in evaluations:

            claim = evaluation["claim_id"]

            for metric in evaluation["metrics"]:

                ws.append([

                    claim,

                    metric.metric_name,

                    metric.score,

                    metric.threshold,

                    metric.passed,

                    metric.remarks

                ])

    # ------------------------------------------------------

    def build_claim_details(
        self,
        ws,
        evaluations
    ):

        ws.append([

            "Claim ID",

            "Decision",

            "Confidence",

            "Execution Time",

            "Token Usage"

        ])

        self.style_header(ws)

        for e in evaluations:

            result = e["agent_result"]

            ws.append([

                result.claim_id,

                result.decision,

                result.confidence,

                result.execution_time,

                result.token_usage.get(
                    "total_tokens",
                    0
                )

            ])